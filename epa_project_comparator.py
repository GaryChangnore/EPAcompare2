#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EPA 專案版本比對工具
功能：比較多個時間點的 EPA 專案 Excel 檔案，標示實質變動
"""

import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class EPAProjectComparator:
    """EPA 專案版本比對器"""
    
    # 不進行比對的欄位（即使不同也不標色）
    EXCLUDED_COLUMNS = {'Seq', 'Snapshot_Date', 'Comments', 'Notes', '備註', '註記'}
    
    # 專案比對 key 欄位（依優先順序）
    PROJECT_KEY_COLUMNS = ['Project Name', 'Applicant Name', '專案名稱', '申請人名稱']
    
    # 顏色定義
    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 🟡 黃色
    RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')     # 🔴 紅色
    
    def __init__(self, excel_files: List[str], snapshot_dates: Optional[Dict[str, str]] = None):
        """
        初始化比對器
        
        Args:
            excel_files: Excel 檔案路徑列表
            snapshot_dates: 可選，手動指定檔案對應的日期 {檔案路徑: 'YYYY/MM/DD'}
        """
        self.excel_files = excel_files
        self.snapshot_dates = snapshot_dates or {}
        self.dataframes = []
        self.file_metadata = []
        
    def _get_file_time(self, file_path: str) -> str:
        """
        判斷檔案時間（優先順序：使用者指定 > 檔案修改時間）
        
        Returns:
            YYYY/MM/DD 格式的日期字串
        """
        # 優先使用使用者指定日期
        if file_path in self.snapshot_dates:
            return self.snapshot_dates[file_path]
        
        # 使用檔案修改時間
        file_stat = os.stat(file_path)
        mod_time = datetime.fromtimestamp(file_stat.st_mtime)
        return mod_time.strftime('%Y/%m/%d')
    
    def _load_excel_files(self) -> None:
        """載入所有 Excel 檔案並進行前處理"""
        for idx, file_path in enumerate(self.excel_files, start=1):
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"檔案不存在: {file_path}")
            
            # 讀取 Excel
            df = pd.read_excel(file_path)
            
            # 判斷時間
            snapshot_date = self._get_file_time(file_path)
            
            # 新增 Seq 和 Snapshot_Date 欄位（放在最前方）
            df.insert(0, 'Snapshot_Date', snapshot_date)
            df.insert(0, 'Seq', idx)
            
            self.dataframes.append(df)
            self.file_metadata.append({
                'file_path': file_path,
                'snapshot_date': snapshot_date,
                'seq': idx,
                'columns': list(df.columns)
            })
    
    def _check_column_structure(self) -> Dict[str, bool]:
        """
        檢查所有檔案的欄位結構是否一致
        
        Returns:
            {專案key: 是否結構異常} 的字典
        """
        if len(self.dataframes) < 2:
            return {}
        
        # 以第一個檔案為基準
        base_columns = self.file_metadata[0]['columns']
        base_col_set = set(base_columns)
        
        structure_issues = {}
        
        # 檢查每個檔案
        for metadata in self.file_metadata[1:]:
            current_columns = metadata['columns']
            current_col_set = set(current_columns)
            
            # 檢查欄位數量、名稱、順序
            if (len(current_columns) != len(base_columns) or
                current_col_set != base_col_set or
                current_columns != base_columns):
                # 結構不一致，標記所有專案
                structure_issues['__ALL__'] = True
                break
        
        return structure_issues
    
    def _find_project_key_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        尋找專案比對 key 欄位
        
        Returns:
            欄位名稱，若找不到則返回 None
        """
        for key_col in self.PROJECT_KEY_COLUMNS:
            if key_col in df.columns:
                return key_col
        return None
    
    def _normalize_key(self, value: str) -> str:
        """
        正規化專案 key（去除前後空白、轉小寫）
        
        Args:
            value: 原始值
            
        Returns:
            正規化後的字串
        """
        if pd.isna(value):
            return ''
        return str(value).strip().lower()
    
    def _merge_projects(self) -> pd.DataFrame:
        """
        合併所有時間點的專案資料
        
        Returns:
            合併後的 DataFrame，依 Snapshot_Date 排序
        """
        # 檢查欄位結構
        structure_issues = self._check_column_structure()
        has_structure_issue = '__ALL__' in structure_issues
        
        if has_structure_issue:
            # 結構異常，只合併最新時間點的資料並標記
            latest_df = self.dataframes[-1].copy()
            latest_df['__STRUCTURE_ERROR__'] = True
            # 為了後續處理，需要建立 __NORMALIZED_KEY__ 欄位
            key_column = self._find_project_key_column(latest_df)
            if key_column:
                latest_df['__NORMALIZED_KEY__'] = latest_df[key_column].apply(self._normalize_key)
            else:
                latest_df['__NORMALIZED_KEY__'] = ''
            return latest_df
        
        # 合併所有資料
        merged_df = pd.concat(self.dataframes, ignore_index=True)
        
        # 找出專案 key 欄位
        key_column = self._find_project_key_column(merged_df)
        if key_column is None:
            raise ValueError("找不到專案比對欄位（Project Name 或 Applicant Name）")
        
        # 正規化 key
        merged_df['__NORMALIZED_KEY__'] = merged_df[key_column].apply(self._normalize_key)
        
        # 依 Snapshot_Date 和 Seq 排序（舊 → 新）
        merged_df = merged_df.sort_values(['__NORMALIZED_KEY__', 'Snapshot_Date', 'Seq'], 
                                          ascending=[True, True, True])
        
        return merged_df
    
    def _compare_fields(self, merged_df: pd.DataFrame) -> pd.DataFrame:
        """
        比對欄位並標記變動
        
        Args:
            merged_df: 合併後的 DataFrame
            
        Returns:
            新增了變動標記的 DataFrame
        """
        # 檢查是否有結構錯誤
        if '__STRUCTURE_ERROR__' in merged_df.columns:
            merged_df['__HAS_CHANGE__'] = False
            merged_df['__CHANGED_CELLS__'] = None
            return merged_df
        
        # 找出專案 key 欄位
        key_column = self._find_project_key_column(merged_df)
        
        # 取得所有欄位（排除不比較的欄位）
        all_columns = [col for col in merged_df.columns 
                      if col not in self.EXCLUDED_COLUMNS and 
                      not col.startswith('__')]
        
        # 初始化標記欄位
        merged_df['__HAS_CHANGE__'] = False
        merged_df['__CHANGED_CELLS__'] = None
        
        # 依專案分組比對
        for key_value in merged_df['__NORMALIZED_KEY__'].unique():
            if not key_value:  # 跳過空值
                continue
            
            project_rows = merged_df[merged_df['__NORMALIZED_KEY__'] == key_value].copy()
            
            if len(project_rows) < 2:
                # 只有一個時間點，無需比對
                continue
            
            # 只比較最新時間點與前一個時間點
            project_rows = project_rows.sort_values(['Snapshot_Date', 'Seq'], ascending=[True, True])
            latest_idx = project_rows.index[-1]
            previous_idx = project_rows.index[-2]
            
            latest_row = project_rows.loc[latest_idx]
            previous_row = project_rows.loc[previous_idx]
            
            # 比對每個欄位
            changed_cells = []
            for col in all_columns:
                if col not in latest_row.index or col not in previous_row.index:
                    continue
                
                latest_val = latest_row[col]
                previous_val = previous_row[col]
                
                # 比較值（處理 NaN）
                if pd.isna(latest_val) and pd.isna(previous_val):
                    continue
                elif pd.isna(latest_val) or pd.isna(previous_val):
                    changed_cells.append(col)
                elif str(latest_val).strip() != str(previous_val).strip():
                    changed_cells.append(col)
            
            # 標記變動
            if changed_cells:
                merged_df.loc[latest_idx, '__HAS_CHANGE__'] = True
                merged_df.loc[latest_idx, '__CHANGED_CELLS__'] = ','.join(changed_cells)
        
        return merged_df
    
    def _apply_colors_to_excel(self, output_path: str, merged_df: pd.DataFrame) -> None:
        """
        將顏色標記應用到 Excel 檔案
        
        Args:
            output_path: 輸出檔案路徑
            merged_df: 已標記變動的 DataFrame
        """
        # 先寫入 Excel（不含顏色）
        merged_df_clean = merged_df.drop(columns=[col for col in merged_df.columns if col.startswith('__')])
        merged_df_clean.to_excel(output_path, index=False, engine='openpyxl')
        
        # 使用 openpyxl 添加顏色
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # 取得欄位名稱對應的欄位索引
        header_row = 1
        column_map = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            column_map[cell.value] = col_idx
        
        # 找出專案 key 欄位
        key_column = self._find_project_key_column(merged_df_clean)
        key_col_idx = column_map.get(key_column)
        seq_col_idx = column_map.get('Seq')
        date_col_idx = column_map.get('Snapshot_Date')
        
        # 檢查是否有結構錯誤
        has_structure_error = '__STRUCTURE_ERROR__' in merged_df.columns
        
        # 建立 DataFrame 索引到 Excel 行號的對應表
        # 因為 to_excel(index=False)，所以 Excel 行號 = DataFrame 位置 + 2（標題行 + 1-based）
        index_to_excel_row = {}
        for pos, df_idx in enumerate(merged_df.index):
            index_to_excel_row[df_idx] = pos + 2  # +1 for header, +1 for 0-based to 1-based
        
        if has_structure_error:
            # 結構異常：標記所有列為紅色（因為結構不一致，無法比對）
            for df_idx in merged_df.index:
                excel_row = index_to_excel_row[df_idx]
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = self.RED_FILL
        else:
            # 依專案分組處理
            for key_value in merged_df['__NORMALIZED_KEY__'].unique():
                if not key_value:
                    continue
                
                project_rows = merged_df[merged_df['__NORMALIZED_KEY__'] == key_value]
                project_rows = project_rows.sort_values(['Snapshot_Date', 'Seq'], ascending=[True, True])
                
                # 取得 Excel 中的行號
                excel_row_indices = [index_to_excel_row[df_idx] for df_idx in project_rows.index]
                
                # 檢查變動
                latest_df_idx = project_rows.index[-1]
                if merged_df.loc[latest_df_idx, '__HAS_CHANGE__']:
                    latest_excel_row = excel_row_indices[-1]
                    changed_cells_str = merged_df.loc[latest_df_idx, '__CHANGED_CELLS__']
                    changed_columns = changed_cells_str.split(',') if pd.notna(changed_cells_str) else []
                    
                    # 標記變動的儲存格為黃色
                    for col_name in changed_columns:
                        if col_name in column_map:
                            col_idx = column_map[col_name]
                            ws.cell(row=latest_excel_row, column=col_idx).fill = self.YELLOW_FILL
                    
                    # 同時標記 Seq、Snapshot_Date、專案名稱欄位為黃色
                    if seq_col_idx:
                        ws.cell(row=latest_excel_row, column=seq_col_idx).fill = self.YELLOW_FILL
                    if date_col_idx:
                        ws.cell(row=latest_excel_row, column=date_col_idx).fill = self.YELLOW_FILL
                    if key_col_idx:
                        ws.cell(row=latest_excel_row, column=key_col_idx).fill = self.YELLOW_FILL
            else:
                # 檢查變動
                latest_df_idx = project_rows.index[-1]
                if merged_df.loc[latest_df_idx, '__HAS_CHANGE__']:
                    latest_excel_row = excel_row_indices[-1]
                    changed_cells_str = merged_df.loc[latest_df_idx, '__CHANGED_CELLS__']
                    changed_columns = changed_cells_str.split(',') if pd.notna(changed_cells_str) else []
                    
                    # 標記變動的儲存格為黃色
                    for col_name in changed_columns:
                        if col_name in column_map:
                            col_idx = column_map[col_name]
                            ws.cell(row=latest_excel_row, column=col_idx).fill = self.YELLOW_FILL
                    
                    # 同時標記 Seq、Snapshot_Date、專案名稱欄位為黃色
                    if seq_col_idx:
                        ws.cell(row=latest_excel_row, column=seq_col_idx).fill = self.YELLOW_FILL
                    if date_col_idx:
                        ws.cell(row=latest_excel_row, column=date_col_idx).fill = self.YELLOW_FILL
                    if key_col_idx:
                        ws.cell(row=latest_excel_row, column=key_col_idx).fill = self.YELLOW_FILL
        
        # 儲存檔案
        wb.save(output_path)
    
    def compare_and_export(self, output_path: str) -> str:
        """
        執行完整比對流程並匯出結果
        
        Args:
            output_path: 輸出 Excel 檔案路徑
            
        Returns:
            輸出檔案路徑
        """
        print("📂 開始載入 Excel 檔案...")
        self._load_excel_files()
        print(f"✅ 已載入 {len(self.dataframes)} 個檔案")
        
        print("🔍 檢查欄位結構...")
        structure_issues = self._check_column_structure()
        if structure_issues:
            print("⚠️  警告：發現欄位結構不一致！")
        else:
            print("✅ 欄位結構檢查通過")
        
        print("🔗 合併專案資料...")
        merged_df = self._merge_projects()
        print(f"✅ 已合併 {len(merged_df)} 筆資料")
        
        print("🔎 比對欄位變動...")
        merged_df = self._compare_fields(merged_df)
        changed_count = merged_df['__HAS_CHANGE__'].sum()
        print(f"✅ 發現 {changed_count} 筆專案有變動")
        
        print("🎨 套用顏色標記...")
        self._apply_colors_to_excel(output_path, merged_df)
        print(f"✅ 結果已匯出至: {output_path}")
        
        return output_path


def main():
    """主程式入口（範例使用）"""
    import sys
    
    if len(sys.argv) < 3:
        print("使用方法:")
        print("  python epa_project_comparator.py <輸出檔案> <檔案1> [檔案2] [檔案3] ...")
        print("\n範例:")
        print("  python epa_project_comparator.py output.xlsx file1.xlsx file2.xlsx file3.xlsx")
        print("\n可選：手動指定日期（使用 --date 參數）")
        print("  python epa_project_comparator.py output.xlsx file1.xlsx --date file1.xlsx:2024/01/15 file2.xlsx --date file2.xlsx:2024/02/20")
        sys.exit(1)
    
    output_path = sys.argv[1]
    excel_files = []
    snapshot_dates = {}
    
    # 解析參數
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '--date' and i + 1 < len(sys.argv):
            date_spec = sys.argv[i + 1]
            if ':' in date_spec:
                file_path, date_str = date_spec.split(':', 1)
                snapshot_dates[file_path] = date_str
                i += 2
            else:
                i += 1
        else:
            if arg.endswith('.xlsx') or arg.endswith('.xls'):
                excel_files.append(arg)
            i += 1
    
    if len(excel_files) < 2:
        print("❌ 錯誤：至少需要 2 個 Excel 檔案")
        sys.exit(1)
    
    # 執行比對
    comparator = EPAProjectComparator(excel_files, snapshot_dates)
    comparator.compare_and_export(output_path)


if __name__ == '__main__':
    main()
